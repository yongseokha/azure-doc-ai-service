import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.terms import TermsVerificationResult

ITEM_COLUMNS = ["itemNm", "value", "subClaim", "evidence", "page", "article", "reason", "status"]
STATUS_ORDER = ["MATCHED", "PARTIAL_MATCH", "MISMATCH"]
STATUS_LABELS = {"MATCHED": "일치", "PARTIAL_MATCH": "부분 일치", "MISMATCH": "불일치"}
HEADER_FILL_COLOR = "0000A5"
HEADER_FONT_COLOR = "FFFFFF"
STATUS_STYLE = {
    "일치": {"fill": "C6EFCE", "font": "006100"},
    "부분 일치": {"fill": "FFEB9C", "font": "9C6500"},
    "불일치": {"fill": "FFC7CE", "font": "9C0006"},
    "약관 검증 필요": {"fill": "FFF9C4", "font": "CC0000"},
    "약관 검증 패스": {"fill": "FFF9C4", "font": "1B7A1B"},
}
COLUMN_LABELS = {
    "name": "대상명",
    "termNm": "약관명",
    "itemNm": "지식 항목",
    "value": "상품 지식",
    "subClaim": "상품 지식 단위",
    "evidence": "약관 내 해당 문구",
    "page": "약관 내 페이지",
    "article": "조항",
    "reason": "차이 설명",
    "status": "검토 결과",
    "totalCnt": "총 검증 항목",
    "matchedCnt": "일치",
    "partialMatchCnt": "부분 일치",
    "mismatchCnt": "불일치",
    "overallResult": "종합 판정",
}
# 왼쪽+위 정렬(줄바꿈): 문장 길이가 제각각이라 여러 줄로 넘어갈 수 있는 컬럼들
TOP_LEFT_COLUMNS = {"itemNm", "value", "subClaim", "evidence", "reason"}
# 중앙+수직중앙 정렬: 짧은 값(숫자/코드성 텍스트) 컬럼들
CENTER_COLUMNS = {
    "page", "article",
    "totalCnt", "matchedCnt", "partialMatchCnt", "mismatchCnt",
    "일치", "부분 일치", "불일치", "합계",
}
# 왼쪽+수직중앙 정렬: 짧은 식별용 텍스트(상품명/문서명) 컬럼들
LEFT_CENTER_COLUMNS = {"name", "termNm"}

COLUMN_WIDTHS = [22, 18, 18, 38, 12, 16, 32, 14]  # itemNm,value,subClaim,evidence,page,article,reason,status
SUMMARY_COLUMN_WIDTHS = [22, 16, 16, 16, 16, 24]  # name,totalCnt,matchedCnt,partialMatchCnt,mismatchCnt,overallResult
MAX_COL = len(COLUMN_WIDTHS)
SUMMARY_MAX_COL = len(SUMMARY_COLUMN_WIDTHS)

_THIN_SIDE = Side(style="thin", color="B7B7B7")
BOX_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

TITLE_FILL = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
BANNER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BANNER_FONT = Font(bold=True, size=13, color="375623")
STATS_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _flatten(result: TermsVerificationResult) -> pd.DataFrame:
    rows = [
        {
            "name": name_result.name,
            "termNm": doc_result.termNm,
            **{col: getattr(item, col) for col in ITEM_COLUMNS},
        }
        for name_result in result.data
        for doc_result in name_result.documents
        for item in doc_result.items
    ]
    return pd.DataFrame(rows, columns=["name", "termNm", *ITEM_COLUMNS])


def _grouped_stats(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=[*group_cols, *STATUS_LABELS.values(), "합계"])
    pivot = (
        df.groupby(group_cols, sort=False)["status"]
        .value_counts()
        .unstack(fill_value=0)
        .reindex(columns=STATUS_ORDER, fill_value=0)
    )
    pivot["합계"] = pivot.sum(axis=1)
    pivot = pivot.rename(columns=STATUS_LABELS)
    return pivot.reset_index()


def _build_item_row(item) -> dict:
    row = {col: getattr(item, col) for col in ITEM_COLUMNS}
    row["status"] = STATUS_LABELS.get(row["status"], row["status"])
    return row


def _stats_from_items(items) -> dict:
    counts = dict.fromkeys(STATUS_ORDER, 0)
    for item in items:
        counts[item.status] = counts.get(item.status, 0) + 1
    total = sum(counts.values())
    overall_result = "약관 검증 필요" if counts["MISMATCH"] > 0 else "약관 검증 패스"
    return {
        "totalCnt": total,
        "matchedCnt": counts["MATCHED"],
        "partialMatchCnt": counts["PARTIAL_MATCH"],
        "mismatchCnt": counts["MISMATCH"],
        "overallResult": overall_result,
    }


def compute_overview_stats(result: TermsVerificationResult) -> dict:
    """전체 항목 수, status별 건수, 종합 판정을 계산한다.

    엑셀 리포트와 콜백의 vrfDataResltJson이 같은 숫자를 쓰도록 양쪽에서 공유하는 함수다.
    """
    all_items = [item for name_result in result.data for doc_result in name_result.documents for item in doc_result.items]
    return _stats_from_items(all_items)


def build_result_payload(result: TermsVerificationResult, verified_at: str) -> dict:
    """콜백의 vrfDataResltJson과 GET 상태조회가 공유하는 전체 결과 payload를 만든다.

    통계(totalCnt/matchedCnt/partialMatchCnt/mismatchCnt/overallResult)를 전체/name별/
    문서별 세 계층에 각각 인라인으로 넣는다 - 세 계층 다 같은 필드명을 쓰므로 파싱 로직을
    재사용할 수 있다.
    """
    data = []
    for name_result in result.data:
        name_items = [item for doc_result in name_result.documents for item in doc_result.items]
        documents = []
        for doc_result in name_result.documents:
            documents.append(
                {
                    "ocrResltKey": doc_result.ocrResltKey,
                    "termNm": doc_result.termNm,
                    "aplyDate": doc_result.aplyDate,
                    **_stats_from_items(doc_result.items),
                    "items": [item.model_dump() for item in doc_result.items],
                }
            )
        data.append({"name": name_result.name, **_stats_from_items(name_items), "documents": documents})

    return {
        "knwlgNm": result.knwlgNm,
        "verifiedAt": verified_at,
        **compute_overview_stats(result),
        "data": data,
    }


def _write_title(ws, row: int, max_col: int) -> int:
    ws.cell(row=row, column=1, value="상품지식-약관 검증 AI 결과 보고서")
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = TITLE_FILL
        cell.border = BOX_BORDER
    title_cell = ws.cell(row=row, column=1)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    ws.row_dimensions[row].height = 28
    return row + 2


def _write_section_title(ws, row: int, title: str, max_col: int) -> int:
    ws.cell(row=row, column=1, value=title)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = BANNER_FILL
        cell.border = BOX_BORDER
    ws.cell(row=row, column=1).font = BANNER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    return row + 1


def _run_ids(keys: list) -> list[int]:
    """keys가 바로 앞과 같으면 같은 run으로 묶는다. 값 하나만으로는 판단하지 않고, 여러 병합
    대상 컬럼을 하나의 튜플 키로 합쳐서 넘겨야 한다 - 그래야 예를 들어 value가 둘 다 None인
    서로 다른 항목이 같은 값으로 오인되어 잘못 병합되는 걸 막을 수 있다."""
    ids = []
    current_id = 0
    for i, key in enumerate(keys):
        if i > 0 and key != keys[i - 1]:
            current_id += 1
        ids.append(current_id)
    return ids


def _merge_consecutive_rows(ws, col_idx: int, start_row: int, run_ids: list[int]) -> None:
    """run_ids가 연속으로 같은 구간을 세로로 셀병합한다 (claim 분해로 반복되는 값 정리용)."""
    run_start = 0
    n = len(run_ids)
    for i in range(1, n + 1):
        if i == n or run_ids[i] != run_ids[run_start]:
            if i - run_start > 1:
                ws.merge_cells(
                    start_row=start_row + run_start, start_column=col_idx, end_row=start_row + i - 1, end_column=col_idx
                )
                ws.cell(row=start_row + run_start, column=col_idx).alignment = Alignment(vertical="top")
            run_start = i


def _write_table(
    ws,
    df: pd.DataFrame,
    start_row: int,
    max_col: int,
    title: str | None = None,
    merge_columns: frozenset = frozenset(),
    highlight_header: bool = False,
    status_column: str | None = None,
    bold_columns: frozenset = frozenset(),
    row_height: int | None = None,
) -> int:
    row = start_row
    if title:
        title_cell = ws.cell(row=row, column=1, value=title)
        title_cell.font = Font(bold=True, size=12)
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row, column=col_idx).border = BOX_BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        row += 1

    if df.empty:
        ws.cell(row=row, column=1, value="(데이터 없음)").border = BOX_BORDER
        return row + 2

    header_font = Font(bold=True, color=HEADER_FONT_COLOR) if highlight_header else Font(bold=True)
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    for col_idx, col_name in enumerate(df.columns, start=1):
        header = COLUMN_LABELS.get(col_name, str(col_name))
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.border = BOX_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill if highlight_header else STATS_HEADER_FILL
    row += 1
    data_start_row = row

    for row_idx, (_, record) in enumerate(df.iterrows()):
        zebra = row_idx % 2 == 1
        if row_height:
            ws.row_dimensions[row].height = row_height
        for col_idx, col_name in enumerate(df.columns, start=1):
            value = record[col_name]
            cell = ws.cell(row=row, column=col_idx, value=None if pd.isna(value) else value)
            cell.border = BOX_BORDER

            if col_name in TOP_LEFT_COLUMNS:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif col_name in CENTER_COLUMNS:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in LEFT_CENTER_COLUMNS:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            if col_name == status_column and value in STATUS_STYLE:
                style = STATUS_STYLE[value]
                cell.fill = PatternFill(start_color=style["fill"], end_color=style["fill"], fill_type="solid")
                cell.font = Font(bold=True, color=style["font"])
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in bold_columns:
                cell.font = Font(bold=True)
            elif zebra:
                cell.fill = ZEBRA_FILL
        row += 1

    active_merge_cols = [col for col in df.columns if col in merge_columns]
    if active_merge_cols:
        keys = list(zip(*(df[col].tolist() for col in active_merge_cols)))
        run_ids = _run_ids(keys)
        for col_name in active_merge_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            _merge_consecutive_rows(ws, col_idx, data_start_row, run_ids)

    return row + 1


def _write_verification_info(ws, row: int, max_col: int, verified_at: str, result: TermsVerificationResult) -> int:
    row = _write_section_title(ws, row, "검증 정보", max_col)

    all_items = [item for name_result in result.data for doc_result in name_result.documents for item in doc_result.items]
    keywords = list(dict.fromkeys(item.itemNm for item in all_items))

    unique_docs = list({doc_result.ocrResltKey: doc_result for name_result in result.data for doc_result in name_result.documents}.values())
    term_list = [
        f"{doc_result.termNm}({doc_result.aplyDate})" if doc_result.aplyDate else doc_result.termNm
        for doc_result in unique_docs
    ]

    label_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    for label, value in [
        ("지식명", result.knwlgNm or ""),
        ("검증 일시", verified_at),
        ("검색 키워드", ", ".join(keywords)),
        ("약관 시행일", ", ".join(term_list)),
    ]:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.fill = label_fill
        label_cell.border = BOX_BORDER
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.border = BOX_BORDER
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max_col)
        for col_idx in range(3, max_col + 1):
            ws.cell(row=row, column=col_idx).border = BOX_BORDER
        row += 1
    return row + 1


def _write_result_summary(ws, row: int, max_col: int, result: TermsVerificationResult) -> int:
    row = _write_section_title(ws, row, "검증 결과 요약", max_col)

    records = []
    for name_result in result.data:
        items = [item for doc_result in name_result.documents for item in doc_result.items]
        records.append({"name": name_result.name, **_stats_from_items(items)})
    df = pd.DataFrame(
        records, columns=["name", "totalCnt", "matchedCnt", "partialMatchCnt", "mismatchCnt", "overallResult"]
    )

    return _write_table(
        ws, df, row, max_col, None,
        highlight_header=True, status_column="overallResult", bold_columns=frozenset({"name"}),
        row_height=22,
    )


def _write_detailed_stats(ws, df: pd.DataFrame, start_row: int, max_col: int) -> int:
    row = _write_section_title(ws, start_row, "상세 통계", max_col)
    row = _write_table(ws, _grouped_stats(df, ["name"]), row, max_col, "상품명별 통계", bold_columns=frozenset({"name"}))
    row = _write_table(
        ws, _grouped_stats(df, ["name", "termNm"]), row, max_col, "문서별 통계", bold_columns=frozenset({"name"})
    )
    return row


def _write_item_detail(ws, result: TermsVerificationResult, start_row: int, max_col: int) -> int:
    row = _write_section_title(ws, start_row, "항목별 상세 비교 결과", max_col)
    for name_result in result.data:
        for doc_result in name_result.documents:
            item_df = pd.DataFrame(
                [_build_item_row(item) for item in doc_result.items],
                columns=ITEM_COLUMNS,
            )
            row = _write_table(
                ws,
                item_df,
                row,
                max_col,
                f"{name_result.name} - {doc_result.termNm}",
                merge_columns=frozenset({"itemNm", "value"}),
                highlight_header=True,
                status_column="status",
                bold_columns=frozenset({"itemNm"}),
            )
    return row


def build_report_xlsx(result: TermsVerificationResult, verified_at: str | None = None) -> bytes:
    """검증 결과를 2개 시트로 담은 xlsx를 만든다.

    - 시트1 "검증 요약": 검증 정보(지식명/검증 일시/검색 키워드/약관 시행일) + 검증 결과 요약(name별 통계)
    - 시트2 "상세 결과": 상세 통계(상품명별/문서별) + 항목별 상세 비교 결과

    verified_at을 안 넘기면 호출 시점 시각을 쓴다 - 콜백 JSON(vrfDataResltJson)의
    verifiedAt과 같은 값을 쓰고 싶으면 호출하는 쪽에서 계산해서 넘겨야 한다.
    """
    verified_at = verified_at or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "검증 요약"
    row = 1
    row = _write_title(ws1, row, SUMMARY_MAX_COL)
    row = _write_verification_info(ws1, row, SUMMARY_MAX_COL, verified_at, result)
    _write_result_summary(ws1, row, SUMMARY_MAX_COL, result)
    for col_idx, width in enumerate(SUMMARY_COLUMN_WIDTHS, start=1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    ws2 = wb.create_sheet("상세 결과")
    df = _flatten(result)
    row = 1
    row = _write_detailed_stats(ws2, df, row, MAX_COL)
    _write_item_detail(ws2, result, row, MAX_COL)
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
